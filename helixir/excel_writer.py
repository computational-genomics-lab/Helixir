"""
Helixir Excel output – 3 focused sheets.

Sheet 1 · Summary
    One row per sequence: SS content (all 3 methods) + AA class breakdown.

Sheet 2 · Structure & Composition
    Per-sequence block:
      • Primary sequence (60-char chunks)
      • CF / GOR IV / Consensus prediction strings
      • SS content table (H / E / C counts + %)
      • Structural segments table
      • Amino acid composition + residue class breakdown

Sheet 3 · Residue Map
    Full per-residue table for all sequences:
    Pos · AA · Class · CF · GOR · Consensus · Pα · Pβ · Pt
"""
from __future__ import annotations
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from helixir.parameters import (
    CHOU_FASMAN, AA_NAMES,
    HYDROPHOBIC, POLAR, CHARGED_POS, CHARGED_NEG, AROMATIC, SPECIAL,
)

# ── Palette ───────────────────────────────────────────────────────────────────
_C = {
    "hdr_bg" : "1F4E79", "hdr_fg" : "FFFFFF",
    "sec_bg" : "2E75B6", "sec_fg" : "FFFFFF",
    "sub_bg" : "D6E4F0", "sub_fg" : "1F4E79",
    "helix"  : "FADADD", "strand" : "D6EAF8", "coil"   : "D5F5E3",
    "alt"    : "F2F7FB", "border" : "BFBFBF",
    "red"    : "C00000", "green"  : "1E8449", "blue"   : "1F4E79",
}

_AA_CLASS_COLOR = {
    "Hydrophobic" : "FFF3CD",
    "Polar"       : "D1F2EB",
    "Pos Charged" : "D6EAF8",
    "Neg Charged" : "FADADD",
    "Aromatic"    : "E8DAEF",
    "Special"     : "FDEBD0",
    "Other"       : "F2F3F4",
}

_SS_FILL = {
    "H": PatternFill("solid", start_color="FADADD"),
    "E": PatternFill("solid", start_color="D6EAF8"),
    "C": PatternFill("solid", start_color="D5F5E3"),
}

def _f(h):  return PatternFill("solid", start_color=h)
def _bd():
    s = Side(style="thin", color=_C["border"])
    return Border(left=s, right=s, top=s, bottom=s)
def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def _bf(sz=10, bold=False, c="000000"): return Font(name="Arial", size=sz, bold=bold, color=c)
def _mono(sz=9): return Font(name="Courier New", size=sz)

def _hdr(ws, row, vals, bg=None, fg=None, sz=10, height=20):
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", size=sz, bold=True, color=fg or _C["hdr_fg"])
        c.fill      = _f(bg or _C["hdr_bg"])
        c.border    = _bd()
        c.alignment = _ctr()
    ws.row_dimensions[row].height = height

def _sec(ws, row, title, ncols, bg="sec_bg", fg="sec_fg"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font      = Font(name="Arial", size=11, bold=True, color=_C[fg])
    c.fill      = _f(_C[bg])
    c.alignment = _lft()
    c.border    = _bd()
    ws.row_dimensions[row].height = 18

def _sub(ws, row, title, ncols):
    _sec(ws, row, title, ncols, bg="sub_bg", fg="sub_fg")

def _aa_class(aa):
    if aa in HYDROPHOBIC:  return "Hydrophobic"
    if aa in POLAR:        return "Polar"
    if aa in CHARGED_POS:  return "Pos Charged"
    if aa in CHARGED_NEG:  return "Neg Charged"
    if aa in AROMATIC:     return "Aromatic"
    if aa in SPECIAL:      return "Special"
    return "Other"

def _safe(name, n=28):
    return name[:n].replace("/", "_").replace("\\", "_")

def _class_pct(seq, cls_set):
    return round(sum(1 for a in seq if a in cls_set) / len(seq) * 100, 1) if seq else 0.0


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 · SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_summary(wb: Workbook, all_results: list[dict], meta: dict):
    ws = wb.active; ws.title = "Summary"; ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    NCOLS = 19

    # Title
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    t = ws["A1"]
    t.value = "Helixir  ·  Peptide Secondary Structure Report"
    t.font  = Font(name="Arial", size=18, bold=True, color=_C["hdr_fg"])
    t.fill  = _f(_C["hdr_bg"]); t.alignment = _ctr()
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    sub = ws["A2"]
    sub.value = (f"Generated: {meta['ts']}   |   "
                 f"Input: {meta['src']}   |   "
                 f"Sequences analysed: {len(all_results)}")
    sub.font = _bf(sz=9, c="595959")
    sub.fill = _f("EBF5FB"); sub.alignment = _ctr()
    ws.row_dimensions[2].height = 16

    # Group header row
    # Col layout: # | ID | Sequence | Length | CF(3) | GOR(3) | Consensus(5) | Classes(4)
    groups = [
        ("Sequence Info",             1,  4),
        ("SS Content – Chou-Fasman",  5,  7),
        ("SS Content – GOR IV",       8, 10),
        ("SS Content – Consensus",   11, 15),
        ("Residue Class Breakdown %",16, 19),
    ]
    group_colors = ["1F4E79","1A5276","1A5276","1E8449","7D3C98"]
    for (label, c1, c2), gc in zip(groups, group_colors):
        ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)
        cell = ws.cell(row=3, column=c1, value=label)
        cell.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill      = _f(gc); cell.alignment = _ctr(); cell.border = _bd()
    ws.row_dimensions[3].height = 16

    col_headers = [
        "#", "Sequence ID", "Sequence", "Length",
        "Helix %", "Strand %", "Coil %",
        "Helix %", "Strand %", "Coil %",
        "Helix %", "Strand %", "Coil %",
        "Longest\nHelix", "Longest\nStrand",
        "Hydro-\nphobic%", "Polar%", "Charged%", "Aromatic%",
    ]
    _hdr(ws, 4, col_headers, sz=9, height=32)

    for row_idx, res in enumerate(all_results, 1):
        r    = row_idx + 4
        seq  = res["sequence"]
        cf   = res["stats"]["cf"]
        gor  = res["stats"]["gor"]
        con  = res["stats"]["con"]
        segs = con["segments"]
        lh   = max((s["length"] for s in segs if s["type"] == "H"), default=0)
        le   = max((s["length"] for s in segs if s["type"] == "E"), default=0)
        alt  = _f(_C["alt"]) if row_idx % 2 == 0 else None

        charged_pct = round(
            _class_pct(seq, CHARGED_POS) + _class_pct(seq, CHARGED_NEG), 1
        )
        vals = [
            row_idx, res["id"], seq, res["length"],
            cf["pct"]["H"],  cf["pct"]["E"],  cf["pct"]["C"],
            gor["pct"]["H"], gor["pct"]["E"], gor["pct"]["C"],
            con["pct"]["H"], con["pct"]["E"], con["pct"]["C"],
            lh, le,
            _class_pct(seq, HYDROPHOBIC),
            _class_pct(seq, POLAR),
            charged_pct,
            _class_pct(seq, AROMATIC),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = _bd()
            if col == 3:   # Sequence — monospace, left-aligned
                c.font      = Font(name="Courier New", size=8)
                c.alignment = _lft()
            elif col == 2: # ID — left-aligned
                c.font      = _bf(sz=9)
                c.alignment = _lft()
            else:
                c.font      = _bf(sz=9)
                c.alignment = _ctr()
            if alt and col != 3: c.fill = alt

        # SS % colour highlights (cols shifted +1 with new Sequence column)
        for col, clr in [(5,_C["helix"]),(6,_C["strand"]),(7,_C["coil"]),
                         (8,_C["helix"]),(9,_C["strand"]),(10,_C["coil"]),
                         (11,_C["helix"]),(12,_C["strand"]),(13,_C["coil"])]:
            ws.cell(row=r, column=col).fill = _f(clr)

    # Col widths: # | ID | Sequence | Length | CF×3 | GOR×3 | Con×5 | Classes×4
    widths = [4, 22, 52, 7, 9, 9, 8, 9, 9, 8, 9, 9, 8, 9, 9, 12, 8, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 · STRUCTURE & COMPOSITION
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_structure(wb: Workbook, all_results: list[dict]):
    ws = wb.create_sheet("Structure & Composition")
    ws.sheet_view.showGridLines = False
    CHUNK = 60; NCOLS = 8

    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    t = ws["A1"]
    t.value = "Structure & Composition Detail"
    t.font  = Font(name="Arial", size=14, bold=True, color=_C["hdr_fg"])
    t.fill  = _f(_C["hdr_bg"]); t.alignment = _ctr()
    ws.row_dimensions[1].height = 26

    row = 3
    for res in all_results:
        seq = res["sequence"]
        n   = res["length"]

        # Sequence title bar
        _sec(ws, row, f"▶  {res['id']}   ({n} aa)", NCOLS); row += 1

        # ── Primary sequence ──────────────────────────────────────────────────
        _sub(ws, row, "Primary Sequence", NCOLS); row += 1
        for s in range(0, n, CHUNK):
            chunk = seq[s:s+CHUNK]
            lc = ws.cell(row=row, column=1, value=f"{s+1}–{s+len(chunk)}")
            lc.font = _bf(bold=True, c=_C["blue"]); lc.alignment = _ctr()
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOLS)
            c = ws.cell(row=row, column=2, value=chunk)
            c.font = _mono(10); c.alignment = _lft(); row += 1

        # ── SS prediction strings ─────────────────────────────────────────────
        row += 1
        _sub(ws, row, "Secondary Structure Predictions  (H = Helix   E = Strand   C = Coil)", NCOLS)
        row += 1
        for lbl, ss_str in [("Chou-Fasman", res["ss"]["cf"]),
                              ("GOR IV",      res["ss"]["gor"]),
                              ("Consensus",   res["ss"]["con"])]:
            for s in range(0, n, CHUNK):
                chunk = ss_str[s:s+CHUNK]
                lc = ws.cell(row=row, column=1, value=f"{lbl} {s+1}–{s+len(chunk)}")
                lc.font = _bf(bold=True, c=_C["blue"]); lc.alignment = _ctr()
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOLS)
                c = ws.cell(row=row, column=2, value=chunk)
                c.font = _mono(10); c.alignment = _lft(); row += 1
            row += 1

        # ── SS content table ──────────────────────────────────────────────────
        _sub(ws, row, "Secondary Structure Content", NCOLS); row += 1
        _hdr(ws, row,
             ["Method","Helix (n)","Helix (%)","Strand (n)","Strand (%)","Coil (n)","Coil (%)","Longest  H | E"],
             sz=9, height=22); row += 1
        for mk, ml in [("cf","Chou-Fasman"), ("gor","GOR IV"), ("con","Consensus")]:
            st   = res["stats"][mk]
            segs = st["segments"]
            lh   = max((s["length"] for s in segs if s["type"] == "H"), default=0)
            le   = max((s["length"] for s in segs if s["type"] == "E"), default=0)
            vals = [ml,
                    st["counts"]["H"], f"{st['pct']['H']}%",
                    st["counts"]["E"], f"{st['pct']['E']}%",
                    st["counts"]["C"], f"{st['pct']['C']}%",
                    f"Helix: {lh} aa   |   Strand: {le} aa"]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _bf(sz=9); c.border = _bd(); c.alignment = _ctr()
            for col, clr in [(2,_C["helix"]),(3,_C["helix"]),(4,_C["strand"]),
                              (5,_C["strand"]),(6,_C["coil"]),(7,_C["coil"])]:
                ws.cell(row=row, column=col).fill = _f(clr)
            row += 1

        # ── Segment table ─────────────────────────────────────────────────────
        row += 1
        _sub(ws, row, "Consensus Structural Segments", NCOLS); row += 1
        _hdr(ws, row,
             ["Start","End","Length (aa)","Type","Residue Sequence","","",""],
             sz=9, height=20); row += 1
        TL = {"H": "α-Helix", "E": "β-Strand", "C": "Coil/Loop"}
        TC = {"H": _C["helix"], "E": _C["strand"], "C": _C["coil"]}
        for seg in res["stats"]["con"]["segments"]:
            vals = [seg["start"], seg["end"], seg["length"],
                    TL[seg["type"]], seg["seq"], "", "", ""]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _bf(sz=9); c.fill = _f(TC[seg["type"]])
                c.border = _bd()
                c.alignment = _lft() if col == 5 else _ctr()
            row += 1

        # ── AA composition ────────────────────────────────────────────────────
        row += 1
        _sub(ws, row, "Amino Acid Composition", NCOLS); row += 1
        _hdr(ws, row,
             ["AA","Full Name","Class","Count","Percent (%)","","",""],
             sz=9, height=20); row += 1
        aa_sorted = sorted(
            ((aa, seq.count(aa)) for aa in set(seq)),
            key=lambda x: -x[1]
        )
        for aa, cnt in aa_sorted:
            cls  = _aa_class(aa)
            clr  = _AA_CLASS_COLOR.get(cls, "FFFFFF")
            pct  = round(cnt / n * 100, 1)
            vals = [aa, AA_NAMES.get(aa, aa), cls, cnt, f"{pct}%", "", "", ""]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _bf(sz=9); c.fill = _f(clr)
                c.border = _bd()
                c.alignment = _lft() if col == 2 else _ctr()
            row += 1

        # Class breakdown summary (inline, columns 6-8)
        cr = row - len(aa_sorted)
        class_data = [
            ("Hydrophobic",  _class_pct(seq, HYDROPHOBIC)),
            ("Polar",        _class_pct(seq, POLAR)),
            ("Pos Charged",  _class_pct(seq, CHARGED_POS)),
            ("Neg Charged",  _class_pct(seq, CHARGED_NEG)),
            ("Aromatic",     _class_pct(seq, AROMATIC)),
            ("Special",      _class_pct(seq, SPECIAL)),
        ]
        hc = ws.cell(row=cr, column=6, value="Residue Class")
        hc.font = _bf(sz=9, bold=True)
        hp = ws.cell(row=cr, column=7, value="% of sequence")
        hp.font = _bf(sz=9, bold=True)
        cr += 1
        for cls_name, cls_pct in class_data:
            clr = _AA_CLASS_COLOR.get(cls_name, "FFFFFF")
            lc  = ws.cell(row=cr, column=6, value=cls_name)
            lc.font = _bf(sz=9); lc.fill = _f(clr); lc.border = _bd(); lc.alignment = _lft()
            pc  = ws.cell(row=cr, column=7, value=f"{cls_pct}%")
            pc.font = _bf(sz=9); pc.fill = _f(clr); pc.border = _bd(); pc.alignment = _ctr()
            cr += 1

        row += 2   # spacer

    for i, w in enumerate([14, 7, 7, 12, 10, 16, 10, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 · RESIDUE MAP
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_resmap(wb: Workbook, all_results: list[dict]):
    ws = wb.create_sheet("Residue Map")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "Per-Residue Analysis  (all sequences)"
    t.font  = Font(name="Arial", size=13, bold=True, color=_C["hdr_fg"])
    t.fill  = _f(_C["hdr_bg"]); t.alignment = _ctr()
    ws.row_dimensions[1].height = 24

    _hdr(ws, 2,
         ["Seq ID", "Pos", "AA", "Class",
          "Chou-Fasman", "GOR IV", "Consensus",
          "Pα", "Pβ", "Pt"],
         sz=9, height=24)

    TL  = {"H": "Helix", "E": "Strand", "C": "Coil"}
    row = 3

    for res in all_results:
        seq = res["sequence"]
        cf  = res["ss"]["cf"]
        gor = res["ss"]["gor"]
        con = res["ss"]["con"]

        for i, aa in enumerate(seq):
            props = CHOU_FASMAN.get(aa, CHOU_FASMAN["X"])
            cls   = _aa_class(aa)
            alt   = i % 2 == 1
            vals  = [
                res["id"] if i == 0 else "",
                i + 1, aa, cls,
                TL[cf[i]], TL[gor[i]], TL[con[i]],
                props[0], props[1], props[2],
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font   = _mono(9) if col == 3 else _bf(sz=9)
                c.border = _bd()
                c.alignment = _ctr()
                if col == 4:
                    c.fill = _f(_AA_CLASS_COLOR.get(cls, "FFFFFF"))
                elif col in (5, 6, 7):
                    c.fill = _SS_FILL[[cf[i], gor[i], con[i]][col - 5]]
                elif alt:
                    c.fill = _f(_C["alt"])
            row += 1

        ws.row_dimensions[row].height = 8; row += 1   # spacer

    for i, w in enumerate([18, 6, 6, 12, 10, 10, 10, 7, 7, 7], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Public entry-point ────────────────────────────────────────────────────────

def write_xlsx(all_results: list[dict], output_path: str,
               input_file: str = "") -> None:
    """Build and save the Helixir 3-sheet .xlsx report."""
    wb   = Workbook()
    meta = {
        "ts" : datetime.now().strftime("%Y-%m-%d %H:%M"),
        "src": __import__("os").path.basename(input_file) if input_file else "interactive",
    }
    _sheet_summary(wb, all_results, meta)
    _sheet_structure(wb, all_results)
    _sheet_resmap(wb, all_results)
    wb.save(output_path)
